# backlash technologies Employee: Barbara
# where she is paid based on the number of hours and minutes spent in class everyday
# she needs a time tracking program,
# where she enters the date and the time she started teaching,
# and then the date and time when she finished a particular course.
# The program calculates the hours she spent on a course in class.
# It then calculates the amount of money Barbara made teaching a course.
# Note that: Barbara is paid $5 dollars per hour
# if Barbara works from 6PM to 8:30PM on Monday 27th July, she would make 2.5 X 5 =$12.5 dollars.

import datetime
import getpass
from datetime import date

import openpyxl as excel
from openpyxl.utils import get_column_letter


# Helper Function for int input
def _input(message, in_type=int):
    while True:
        try:
            return in_type(input(message))
        except ValueError:
            pass


class Backslash:
    def __init__(self):
        print("== == == Employee Details == == ==")
        self.date = date.today().strftime("%b-%d-%Y")
        self.name = str(input("Enter Your Name:"))
        self.amount = float(_input("Enter Amount paid per hour:"))
        self.hoursSt = datetime.timedelta(hours=_input("Hours Worked Started:"))
        self.minSt = datetime.timedelta(minutes=_input("Mins Worked Started:"))


class Employee(Backslash):
    def __init__(self):
        super().__init__()
        self.hrsend = datetime.timedelta(hours=_input("Hours Worked Ended:"))
        self.minsend = datetime.timedelta(minutes=_input("Minutes Worked Ended:"))
        self.total = (self.hrsend - self.hoursSt) + (self.minsend - self.minSt)
        self.final_amount = float(((self.total.seconds / 3600) * self.amount))
        self.store_list = [[self.date, self.hoursSt + self.minSt, self.hrsend + self.minsend, self.total,
                            '${:,.2f}'.format(self.final_amount)]]
        # self.store_dict ={{self.date:self.store_list}}
        print("== == == Information Successfully Stored == == ==")
        self.buttons()

    def time_ended(self):
        print("Total Time worked is:\t", self.total)
        print("Total Money is\t", '${:,.2f}'.format(self.final_amount), "\n")

    def ShowEmployeeDetails(self):
        print("\t", self.date,
              "\nEmployee Name:\t", self.name,
              "\nTime Started:\t", self.hoursSt + self.minSt,
              "\nTime Ended:\t\t", self.hrsend + self.minsend,
              "\nAmount PerHour:\t", '${:,.2f}'.format(self.amount), "\n")

    def export(self):
        wb = excel.Workbook()
        sheet = wb.active
        sheet.title = "DSC Challenge"

        heading = ["Date", "Time Started", "Time Ended", "Total Time Worked", "Amount Paid"]
        # this prints out the list onto the first range of cells i.e A1:E1
        heading_row = 1
        for item in heading:
            sheet.column_dimensions[get_column_letter(heading_row)].width = 20
            sheet.cell(row=1, column=heading_row, value=item)
            heading_row += 1
            # ======Print Key and Value=====
        column = 1
        row = 2
        # For Loop to access insides of a list within a list
        for a_list in self.store_list:
            for inner_list_index in a_list:
                sheet.column_dimensions[get_column_letter(column)].width = 20
                sheet.cell(row=row, column=column, value=inner_list_index)
                column += 1
            column = 1
            row += 1
        wb.save("C:/Users/" + getpass.getuser() + '/Desktop/Dsc-Challenge.xlsx')
        print("==Excel File Successfully Created==")

    def update(self):
        self.date = date.today().strftime("%b-%d-%Y")
        self.hoursSt = datetime.timedelta(hours=_input("New Hours Worked Started:"))
        self.minSt = datetime.timedelta(minutes=_input("New Mins Worked Started:"))
        self.hrsend = datetime.timedelta(hours=_input("New Hours Worked Ended:"))
        self.minsend = datetime.timedelta(minutes=_input("New Minutes Worked Ended:"))
        new_start = (self.hoursSt + self.minSt)
        new_end = (self.hrsend + self.minsend)
        self.total = (self.hrsend - self.hoursSt) + (self.minsend - self.minSt)
        self.final_amount = float(((self.total.seconds / 3600) * self.amount))

        if any(self.date in cT for cT in self.store_list):
            self.store_list.append(['', new_start, new_end, self.total, '${:,.2f}'.format(self.final_amount)])
        else:
            self.store_list.append([self.date, new_start, new_end, self.total, '${:,.2f}'.format(self.final_amount)])
        print("==Excel Data Updated Successfully==")
        self.export()

    def buttons(self):
        print("\n\t\t== == == OptionsAvailable == == == ")
        button = ''
        while button != "Q":
            print("== 'T' = Time&Money, 'S' = Display Data, 'E' = Export to Excel, 'U' = Update Data, 'Q' = Exit==,")
            button = str.capitalize(input("\nEnter an Option:"))
            if button == 'T':
                print("\t==Amount Gained==")
                self.time_ended()
            elif button == 'S':
                print("\t==Details on Employee==")
                self.ShowEmployeeDetails()
            elif button == 'E':
                print("\t==Exporting to Excel File==")
                self.export()
            elif button == 'U':
                self.update()
                print("\t==New Data Recorded==")
            elif button == 'Q':
                print("Quitting...")
                break


# Create Employee
obj1 = Employee()
